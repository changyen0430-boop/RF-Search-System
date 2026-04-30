import streamlit as st
import requests
import io
import pandas as pd
import os
import re
import webbrowser
from io import BytesIO

# --- 頁面設定 ---
st.set_page_config(page_title="RF Spec Search Pro", layout="wide", page_icon="📡")

# --- [資料加載邏輯] ---
@st.cache_data(ttl=3600)
def load_data():
    file_path = os.path.join(os.getcwd(), 'Final_Summary_Result_Pro.xlsx')
    if os.path.exists(file_path):
        try:
            df = pd.read_excel(file_path)
            from openpyxl import load_workbook
            wb = load_workbook(file_path, data_only=False)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            if "Link to Datasheet" in headers:
                col_idx = headers.index("Link to Datasheet") + 1
                formulas = []
                for row in range(2, ws.max_row + 1):
                    cell_val = ws.cell(row=row, column=col_idx).value
                    v = str(cell_val) if cell_val else ""
                    if "HYPERLINK" in v.upper():
                        match = re.search(r'HYPERLINK\("([^"]+)"', v, re.I)
                        formulas.append(match.group(1) if match else "NA")
                    else:
                        formulas.append(v if v.startswith("http") else "NA")
                df["Link to Datasheet"] = formulas[:len(df)]
            return df.fillna("NA")
        except Exception as e:
            st.error(f"Excel 讀取失敗: {e}")
            return pd.DataFrame()
    return pd.DataFrame()

df = load_data()

# --- 2. 側邊欄過濾面板 ---
st.sidebar.header("🔍 智能過濾面板")
keyword = st.sidebar.text_input("全域搜尋 (空格隔開)", placeholder="例如: AMPAK 6E", key="search_input")

st.sidebar.divider()

# Wi-Fi & BT 過濾
all_wifi = sorted([x for x in df["Feature Support_Wi-Fi"].unique() if x != "NA"]) if not df.empty else []
selected_wifi = st.sidebar.multiselect("Wi-Fi 規格過濾", all_wifi)

all_bt = sorted([x for x in df["Feature Support_BT"].unique() if x != "NA"]) if not df.empty else []
selected_bt = st.sidebar.multiselect("BT 規格過濾", all_bt)

# --- 【修正】MCU 簡化：改為 Checkbox ---
st.sidebar.subheader("💻 運算核心")
mcu_checked = st.sidebar.checkbox("內建 MCU (SoC)")

st.sidebar.divider()

# IoT 與 NFC 過濾器
st.sidebar.subheader("🌐 IoT 協定過濾")
iot_options = st.sidebar.multiselect("選擇 IoT 協定 (AND 模式)", ["Zigbee", "Thread", "Matter"])

# --- 【修正】NFC 過濾器 ---
st.sidebar.subheader("🆔 安全元件支援")
nfc_checked = st.sidebar.checkbox("NFC (Secure Element)")

# --- 核心過濾邏輯執行區 ---
filtered_df = df.copy()

# A. 全域搜尋
if keyword:
    for term in keyword.split():
        filtered_df = filtered_df[filtered_df.apply(lambda r: r.astype(str).str.contains(term, case=False).any(), axis=1)]

# B. Wi-Fi & BT
if selected_wifi:
    filtered_df = filtered_df[filtered_df["Feature Support_Wi-Fi"].isin(selected_wifi)]
if selected_bt:
    filtered_df = filtered_df[filtered_df["Feature Support_BT"].isin(selected_bt)]

# C. 【修正邏輯】MCU 支援：只要不是 NA 且不是空的，視為有支援
if mcu_checked:
    filtered_df = filtered_df[
        (filtered_df["MCU"].astype(str).str.upper() != "NA") & 
        (filtered_df["MCU"].astype(str).str.strip() != "")
    ]

# D. IoT 邏輯過濾 (AND 模式)
if iot_options:
    for option in iot_options:
        col_name = f"Feature Support_{option}"
        if col_name in filtered_df.columns:
            filtered_df = filtered_df[filtered_df[col_name].astype(str).str.contains(f"YES|{option}", case=False, na=False)]

# D. 【暴力修正】NFC 支援過濾：相容 "YES" 或 "NFC" 字眼
if nfc_checked:
    # 只要欄位內容包含 "YES" 或 "NFC" (不限大小寫) 且不是 "NA"，就顯示
    filtered_df = filtered_df[
        (filtered_df["Feature Support_NFC"].astype(str).str.contains("YES|NFC", case=False, na=False)) &
        (filtered_df["Feature Support_NFC"].astype(str).str.upper() != "NA")
    ]

# --- 3. 下載 Excel 功能 (專業美化版) ---
st.sidebar.divider()
if not filtered_df.empty:
    buffer = io.BytesIO()
    
    # 建立 ExcelWriter
    with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
        filtered_df.to_excel(writer, index=False, sheet_name='RF_Spec_Results')
        
        workbook  = writer.book
        worksheet = writer.sheets['RF_Spec_Results']

        # --- [格式定義] ---
        # 1. 標題格式 (深藍底、白字、加粗、框線、置中)
        header_format = workbook.add_format({
            'bold': True,
            'text_wrap': True,
            'valign': 'vcenter',
            'align': 'center',
            'fg_color': '#4472C4', 
            'font_color': 'white',
            'border': 1
        })

        # 2. 內容格式 (框線、垂直置中、水平置中)
        cell_format = workbook.add_format({
            'valign': 'vcenter',
            'align': 'center',
            'border': 1
        })

        # --- [套用美化] ---
        # A. 套用標題格式
        for col_num, value in enumerate(filtered_df.columns.values):
            worksheet.write(0, col_num, value, header_format)
            
            # B. 自動計算並調整欄寬
            # 取 (標題長度) 與 (內容最長字數) 的最大值，並加上緩衝空間
            max_len = max(
                filtered_df[value].astype(str).map(len).max(),  # 內容長度
                len(str(value))                                 # 標題長度
            ) + 2
            worksheet.set_column(col_num, col_num, min(max_len, 50), cell_format) # 最大寬度限制 50

        # C. 凍結首行
        worksheet.freeze_panes(1, 0)

    st.sidebar.download_button(
        label="📥 下載專業美化版 Excel", 
        data=buffer.getvalue(), 
        file_name=f"RF_Spec_Export_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.xlsx", 
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )
else:
    st.sidebar.info("💡 尚無符合條件的資料可供下載")

# --- 4. 主要顯示區域 ---
st.title("📡 RF Spec Search System by RF Tommy")
st.metric("符合結果", len(filtered_df))

event = st.dataframe(
    filtered_df, use_container_width=True, on_select="rerun", selection_mode="single-row",
    column_config={
        "Link to Datasheet": st.column_config.LinkColumn("Datasheet", display_text="Open PDF"),
        "Vendor PN": st.column_config.TextColumn("Vendor PN", width="medium")
    }
)

# --- 5. 點擊列直接開啟 Datasheet ---
st.divider()
if event and "selection" in event and len(event["selection"]["rows"]) > 0:
    selected_row_idx = event["selection"]["rows"][0]
    row_data = filtered_df.iloc[selected_row_idx]
    pdf_link = str(row_data["Link to Datasheet"])
    pn_name = str(row_data.get("Vendor PN", ""))

    if pdf_link != "NA":
        st.success(f"✅ 已選取 {pn_name}。")
        st.components.v1.html(
            f"<script>window.open('{pdf_link}', '_blank');</script>",
            height=0,
        )
        st.link_button(f"🚀 點此開啟 {pn_name} Datasheet", pdf_link, use_container_width=True)
    else:
        st.error("⚠️ 此型號目前無 PDF 連結。")

    with st.expander("🧠 AI 決策邏輯分析", expanded=True):
        log_file = 'ai_debug_report.txt'
        if os.path.exists(log_file):
            try:
                with open(log_file, 'r', encoding='utf-8') as f:
                    all_logs = f.read()
                    search_pattern = rf"=== Logic Analysis for:.*{pn_name}.*===(.*?)(?==== Logic Analysis for:|$)"
                    match = re.search(search_pattern, all_logs, re.DOTALL | re.IGNORECASE)
                    if match: st.code(match.group(0).strip(), language="text")
                    else: st.info(f"💡 找不到 {pn_name} 的日誌。")
            except: st.error("日誌讀取失敗。")